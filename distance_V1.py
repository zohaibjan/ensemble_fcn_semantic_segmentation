from PIL import Image
from scipy import ndimage
import numpy as np
import os
from openpyxl import Workbook, load_workbook
import math
import cv2

def rgb2gray(rgb):
	return np.dot(rgb[...,:3], [0.2989, 0.5870, 0.1140])

def data_loader(image_path, label_path, object_class):
	im = np.array(Image.open(image_path))
	label = np.array(Image.open(label_path))
	idx_bg = label!=object_class
	idx_obj = label==object_class
	im[idx_bg] = [0,0,0]
	return rgb2gray(im)
	
def bottom(im):
	b = 0
	for i in range(im.shape[0]): # height
		if 1 in im[i]:
			b = i
	return b, first_object_range(im[b])
	
def first_object_range(arr):
	flag = True
	update_start = True
	start = 0
	end = 0
	for j in range(arr.shape[0]): # width
		if flag:
			if arr[j] != 0:
				if update_start:
					start = j
					update_start = False
				end = j
			if arr[j] == 0 and not update_start:
				flag = False
	return [start, end]
	
def first_object_range_right(arr, p):
	flag = True
	update_start = True
	start = 0
	end = 0
	for j in range(arr.shape[0]): # width
		if j < p:
			continue
		if flag:
			if arr[j] != 0:
				if update_start:
					start = j
					update_start = False
				end = j
			if arr[j] == 0 and not update_start:
				flag = False
	return [start, end]

def compute_distances(f, frame_path, seg_path, class_attribute, class_line):
	ori_name = frame_path + f
	pred_name = seg_path + f.replace('.jpg', '_pred.png')
	pred = np.array(Image.open(pred_name))
	pred = resize(pred, 1600, 1200)
	
	# class_attribute
	attr_bottoms = []
	attr_ranges = []
	if pred[pred==class_attribute].shape[0] > THRESHOLD_P:
		object_class = class_attribute
		im = data_loader(ori_name, pred_name, object_class)
		im = resize(im, 1600, 1200)
		filtered_im = ndimage.gaussian_filter(im, THRESHOLD_G)
		im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
		#pixel filter
		filtered_count = 0
		### for object
		for i in range(count+1):
			if i == 0:
				continue
			# pixels of the i-th object 
			idx_oi = im_c==i
			if im_c[idx_oi].shape[0] > THRESHOLD_P:
				temp = np.zeros([im.shape[0], im.shape[1]], dtype = np.uint8)
				temp[idx_oi] = 1
				filtered_count += 1
				b, r = bottom(temp)
				attr_bottoms.append(b)
				attr_ranges.append(r)
		'''
		if filtered_count > 0:
			print('Number of class ' + str(object_class) + ' objects is ' + str(filtered_count))
			plt.imshow(im_c)
			plt.show()
		'''
		
	# line as reference
	distances = []
	if pred[pred==class_line].shape[0] > THRESHOLD_P:
		object_class = class_line
		im = data_loader(ori_name, pred_name, object_class)
		im = resize(im, 1600, 1200)
		filtered_im = ndimage.gaussian_filter(im, THRESHOLD_G)
		im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
		
		im_road = data_loader(ori_name, pred_name, 9)
		im_road = resize(im_road, 1600, 1200)
		filtered_im_road = ndimage.gaussian_filter(im_road, THRESHOLD_G)
		im_road_c, road_count = ndimage.label(filtered_im_road > THRESHOLD_C)

		pixel_location = ''
		passenger_side = 0
		driver_side = 0
		for i in range(len(attr_bottoms)):
			line_range = first_object_range(im_c[attr_bottoms[i]])
			#road_range = first_object_range(im_road_c[attr_bottoms[i]])
			road_range = first_object_range_right(im_road_c[attr_bottoms[i]], attr_ranges[i][1])
			pixel_location += '(' + str(attr_bottoms[i]) + ',' + str(attr_ranges[i][1]) + ');'
			if attr_ranges[i][1] <= 800:
				passenger_side = 1
			if attr_ranges[i][1] > 800:
				driver_side = 1
			# if line detected then compute distance
			pixel_size = 2.0
			if line_range[1] != 0:
				line_w = line_range[1] - line_range[0] + 1
				pixel_size = LINE_WIDTH_REF*1.0/line_w
				''' distance to line
				distance = (line_range[0] - attr_ranges[i][1])*pixel_size
				if distance > 0:
					distances.append(distance/100)
				'''
			if road_range[1] != 0:
				distance = (road_range[0] - attr_ranges[i][1])*pixel_size
				if distance >= 0:
					distances.append(distance/100)
	return distances, pixel_location, passenger_side, driver_side

def attribute_list(f, frame_path, seg_path, CLASS_NUMBER):
	print('processing image: ' + f)
	attributes = np.zeros([CLASS_NUMBER], dtype = np.uint8)
	ori_name = frame_path + f
	pred_name = seg_path + f.replace('.jpg', '_pred.png')
	pred = np.array(Image.open(pred_name))
	### obj acc
	for class_i in range(CLASS_NUMBER):
		if class_i != 0:
			if pred[pred==class_i].shape[0] > THRESHOLD_P:
				object_class = class_i
				im = data_loader(ori_name, pred_name, object_class)
				filtered_im = ndimage.gaussian_filter(im, THRESHOLD_G)
				im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
				### for object
				for i in range(count+1):
					if i != 0:
						# pixels of the i-th object 
						idx_oi = im_c==i
						if im_c[idx_oi].shape[0] > THRESHOLD_P:
							### counting
							attributes[class_i] = 1
	return np.delete(attributes, 0)

### meta end
def resize(im, w, h):
	dim = (w, h)
	resized = cv2.resize(im, dim, interpolation = cv2.INTER_AREA)
	if len(resized.shape) > 2 and resized.shape[2] == 4:
		resized = cv2.cvtColor(resized, cv2.COLOR_BGRA2BGR)
	return resized

def compute_distances_meta(f, frame_path, seg_path, class_attribute, class_line):
	pixel_size = math.sqrt( ((3-3)**2)+((10-20)**2) )/math.sqrt( ((1307-1063)**2)+((745-545)**2) )
	ori_name = frame_path + f
	pred_name = seg_path + f.replace('.jpg', '_pred.png')
	pred = np.array(Image.open(pred_name))
	pred = resize(pred, 1600, 1200)
	
	# class_attribute
	attr_bottoms = []
	attr_ranges = []
	if pred[pred==class_attribute].shape[0] > THRESHOLD_P:
		object_class = class_attribute
		im = data_loader(ori_name, pred_name, object_class)
		im = resize(im, 1600, 1200)
		filtered_im = ndimage.gaussian_filter(im, THRESHOLD_G)
		im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
		
		#pixel filter
		filtered_count = 0
		### for object
		for i in range(count+1):
			if i == 0:
				continue
			# pixels of the i-th object 
			idx_oi = im_c==i
			if im_c[idx_oi].shape[0] > THRESHOLD_P:
				temp = np.zeros([im.shape[0], im.shape[1]], dtype = np.uint8)
				temp[idx_oi] = 1
				filtered_count += 1
				b, r = bottom(temp)
				attr_bottoms.append(b)
				attr_ranges.append(r)
		'''
		if filtered_count > 0:
			print('Number of class ' + str(object_class) + ' objects is ' + str(filtered_count))
			plt.imshow(im_c)
			plt.show()
		'''
		
	# meta as reference
	distances = []
	if pred[pred==class_line].shape[0] > THRESHOLD_P:
		object_class = class_line
		im = data_loader(ori_name, pred_name, object_class)
		im = resize(im, 1600, 1200)
		filtered_im = ndimage.gaussian_filter(im, THRESHOLD_G)
		im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
		
		im_road = data_loader(ori_name, pred_name, 9)
		im_road = resize(im_road, 1600, 1200)
		filtered_im_road = ndimage.gaussian_filter(im_road, THRESHOLD_G)
		im_road_c, road_count = ndimage.label(filtered_im_road > THRESHOLD_C)

		for i in range(len(attr_bottoms)):
			road_range = first_object_range_right(im_road_c[attr_bottoms[i]], attr_ranges[i][1])
			# if line detected then compute distance
			if road_range[1] != 0:
				distance = (road_range[0] - attr_ranges[i][1])*pixel_size
				if distance >= 0:
					distances.append(distance)
	return distances
### meta end


def medianType(f,frame_path, seg_path):
	ori_name = frame_path + f
	pred_name = seg_path + f.replace('.jpg', '_pred.png')
	pred = np.array(Image.open(pred_name))
	pred = resize(pred, 1600, 1200)
	medianType={
		"Concrete Barrier": 10,
		"Metal barrier":    6,
		"Median Concrete":  17,
		"Median Grass":     21,
		"Flexipost":        19,
		"Centerline / Wide Centerline":   1
	}
	medianCount={
		"Concrete Barrier": 0,
		"Metal barrier":    0,
		"Median Concrete":  0,
		"Median Grass":     0,
		"Flexipost":        0,
		"Centerline / Wide Centerline":   0
	}

	for key, val in medianType.items():
		if pred[pred==val].shape[0] > THRESHOLD_P:
			image = data_loader(ori_name, pred_name, val)
			image = resize(image, 1600, 1200)
			filtered_im = ndimage.gaussian_filter(image, THRESHOLD_G)
			im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
			medianCount[key] = count
	return medianCount            


def road_angle(f, frame_path, seg_path):
	ori_name = frame_path + f
	pred_name = seg_path + f.replace('.jpg', '_pred.png')
	pred = np.array(Image.open(pred_name))
	pred = resize(pred, 1600, 1200)
	angle = 0
	# road_attribute (label = 9)
	attr_bottoms = []
	attr_ranges = []
	if pred[pred==9].shape[0] > THRESHOLD_P:
		im = data_loader(ori_name, pred_name, 9)
		im = resize(im, 1600, 1200)
		filtered_im = ndimage.gaussian_filter(im, THRESHOLD_G)
		im_c, count = ndimage.label(filtered_im > THRESHOLD_C)
		#pixel filter
		filtered_count = 0
		### for object
		max_count = 0
		max_i = 1
		for i in range(count+1):
			if i == 0:
				continue
			# pixels of the i-th object 
			idx_oi = im_c==i
			if max_count < im_c[idx_oi].shape[0]:
				max_count = im_c[idx_oi].shape[0]
				max_i = i
		idx_oi = im_c==max_i
		temp = np.zeros([im.shape[0], im.shape[1]], dtype = np.uint8)
		temp[idx_oi] = 1
		# fit curve and compute angle
		

	return angle

# other functions
# 1. number of legs of road
# 2. road slope (upwards and downwards)
# 3. roadside severity
# 4. turn lanes (left/right turn lane)
# 5. intersection type
# 6. delineation (centreline, edgeline, guidepost)
# 7. property access point count
# 8. skid resistance/grip (adequate, medium, poor)
# 9. pedestrian crossing marking (on road)
# 10. pedestrian crossing facilities (refuge island, marking, signal)
# 11. parking marking (on road, side)
# 12. sidewalk (sealed, unsealed)
# 13. median type
# 14. land use
# 15. area type



def process_nvgx(path):
	l = []
	with open(path) as nvgx:
		for line in nvgx:
			if ';' in line and 'LEADIN' not in line:
				attrs = line.split(',')
				chainage = attrs[0]
				longitude = attrs[3]
				latitude = attrs[4]
				l.append([chainage,longitude,latitude])
	return l

def generate_excel(nvgx_path, frame_path, seg_path, clazz, output_path):
	gt_file_names = os.listdir(frame_path)

	attributes = np.zeros([clazz], dtype = np.int32)
	output = output_path + 'output.xlsx'
	if os.path.isfile(output):
		os.remove(output)
	cell_titles = ['Image name', 'Chainage', 'Longitude', 'Latitude', 'Lines','Poles','Sign 100','Sign 110','Sign 60','Metal barriers','Trees','Rumble strips','Roads','Concrete barriers','Curvature signs','Guide posts','Merge lanes','Defects','Warning signs','Railway signs','Median concrete','Bicycle paths','Flexiposts','Road work signs','Median grass','Houses and buildings','Signal signs','Pedestrian crossing signs','Pedestrian ccrossings','School zone warning signs','Speed humps','Speed hump signs','Roundabout signs', 'Min distance between Pole and Line (m)', 'Min distance meta_method', 'Object pixel location', 'Pole_passenger_side', 'Pole_driver_side', 'Road_angle'] 

	wb = Workbook()
	ws = wb.active
	ws.append(cell_titles)
	wb.save(output)
	
	chainage = process_nvgx(nvgx_path)

	### for image
	wb = load_workbook(output)
	ws = wb.active
	i = 0
	for f in gt_file_names:
		if '.jpg' not in f:
			print('Input is not an image.')
		else:
			n_list = [f]
			chainage_list = [chainage[i][0]]
			longitude_list = [chainage[i][1]]
			latitude_list = [chainage[i][2]]
            # return 1s for each object detected in the frame
			a_list = attribute_list(f, frame_path, seg_path, clazz)
			try:
				distances, pixel_location, passenger_side, driver_side = compute_distances(f, frame_path, seg_path, 2, 1)
				d_list = [min(distances)]
			except:
				d_list = ['-']
			if d_list == [0]:
				d_list_meta = [0]
			else:
				try:
					d_list_meta = [min(compute_distances_meta(f, frame_path, seg_path, 2, 1))]
				except:
					d_list_meta = ['-']
			# other functions
			medians = medianType(f, frame_path, seg_path)
			angle = road_angle(f, frame_path, seg_path)
			ws.append(n_list + chainage_list + longitude_list + latitude_list + a_list.tolist() + d_list + d_list_meta + [pixel_location] + [passenger_side] + [driver_side] + [angle])
			i += 1
	wb.save(output)

THRESHOLD_G = 0	
THRESHOLD_C = 10
THRESHOLD_P = 50
LINE_WIDTH_REF = 15 # line width 20cm

#generate_excel('10A_1.nvgx', 'output/ori_resized/', 'output/fuse_seg/', 30, 'output/')